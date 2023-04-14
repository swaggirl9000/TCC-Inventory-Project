import tkinter as tk
from tkinter import ttk
import openpyxl #for excel access

#accessing data from excel file 
def load_data():
    workbook = openpyxl.load_workbook("/Users/ada/Desktop/TCC/inventory.xlsx")
    sheet = workbook.active
    values = list(sheet.values)
    
    #header values
    for col in values[0]:
        treeview.heading(col, text=col)
    #vals in excel sheet
    for val in values[1:]:
        treeview.insert("", tk.END, values = val)

def insert():
    #row is inserted into excel
    equipment = equip_entry.get()
    name = name_entry.get()
    stat = status.get()
    reserve = "Reserved" if res.get() else "Unreserved"

    workbook = openpyxl.load_workbook("/Users/ada/Desktop/TCC/inventory.xlsx")
    sheet = workbook.active
    new_data = [equipment, name, stat, reserve]
    sheet.append(new_data)
    workbook.save("/Users/ada/Desktop/TCC/inventory.xlsx")
    #row is inserted into GUI
    treeview.insert("", tk.END, values=new_data)
    equip_entry.delete(0, "end")
    equip_entry.insert(0, "Equipment")
    name_entry.delete(0, "end")
    name_entry.insert(0, "Name")
    status.set(status_types[0])
    res_check.state(["!selected"])

def delete():
    #delete from excel file
    workbook = openpyxl.load_workbook("/Users/ada/Desktop/TCC/inventory.xlsx")
    sheet = workbook.active

    #delete from treeview
    selected = treeview.selection()[0]
    rows = treeview.item(selected)["values"]

    for index, row in enumerate(sheet.values):

        if list(row) == rows:
            sheet.delete_rows(index + 1,1)

    treeview.delete(selected)
    workbook.save("/Users/ada/Desktop/TCC/inventory.xlsx")

root = tk.Tk()

#add theme (this is just to make it pretty)
style = ttk.Style(root)
root.tk.call("source", "forest-dark.tcl")
style.theme_use("forest-dark")

#frame for widgets in order to add data 
frame = ttk.Frame(root)
frame.pack()

#parent widget
widgets_frame = ttk.LabelFrame(frame, text="Insert Data")
widgets_frame.grid(row=0, column=0, padx=30, pady=20)

#child widgets
equip_entry = ttk.Entry(widgets_frame)
equip_entry.insert(0, "Equipment Name")
#delets name of category once user starts typing
equip_entry.bind("<FocusIn>", lambda x: equip_entry.delete("0", "end"))
equip_entry.grid(row=0, column=0, sticky= "ew", padx=5, pady=(0,5))
name_entry = ttk.Entry(widgets_frame)
name_entry.insert(0, "Tutor's Name")
name_entry.bind("<FocusIn>", lambda x: name_entry.delete("0", "end"))
name_entry.grid(row=1, column=0, sticky= "ew", padx=5, pady=(0,5))

#drop down menu for status of equipment
status_types = ["Available", "In Use","Broken", "Reserved for Later", "Other"]
status = ttk.Combobox(widgets_frame, values=status_types)
status.current(0) #default: available
status.grid(row=2, column=0, sticky="ew", padx=5, pady=(0,5))

#reservation option
res = tk.BooleanVar()
res_check = ttk.Checkbutton(widgets_frame, text="Reservation Status", variable=res)
res_check.grid(row=3, column=0, sticky="nsew", padx=5, pady=5)

#insert into excel sheet
button = ttk.Button(widgets_frame, text="Enter Data", command = insert)
button.grid(row=4, column=0, sticky = "nsew", padx=5, pady=(5))

#delete button
del_button = ttk.Button(widgets_frame, text="Delete Data", command = delete)
del_button.grid(row=5, column = 0, sticky="nsew", padx =5, pady=5)

#side frame to display excel file
treeFrame = ttk.Frame(frame)
treeFrame.grid(row=0, column=1, pady=10)
columns = ("Equipment", "Tutor Name", "Status", "Reserved?")
scroll = ttk.Scrollbar(treeFrame)
scroll.pack(side="right", fill="y")
treeview = ttk.Treeview(treeFrame, show="headings", yscrollcommand=scroll.set, columns=columns, height=10)
treeview.column("Equipment", width=150)
treeview.column("Tutor Name", width=150)
treeview.column("Status", width=100)
treeview.column("Reserved?", width=75)
treeview.pack()
scroll.config(command=treeview.yview)
load_data()


root.mainloop()